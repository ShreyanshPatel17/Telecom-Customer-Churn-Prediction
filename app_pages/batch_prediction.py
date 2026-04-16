def show_batch_prediction():

    import streamlit as st
    import pandas as pd
    from io import BytesIO

    from utils.load_models import load_all
    from utils.predict_form import preprocess_input, make_prediction

    # ------------------ LOAD ------------------
    model, scaler, label_encoders, features = load_all()

    st.title("📂 Batch Churn Prediction")
    st.subheader("Upload a CSV file to predict churn for multiple customers", divider=False)


    # ------------------ FILE UPLOAD ------------------
    uploaded_file = st.file_uploader("Upload CSV File", type=["csv"])

    if uploaded_file is not None:

        try:
            df = pd.read_csv(uploaded_file)

            st.subheader("📄 Uploaded Data")
            st.dataframe(df.head())

            # ------------------ REQUIRED COLUMNS ------------------
            required_cols = [
                'gender', 'SeniorCitizen', 'Partner', 'Dependents',
                'tenure', 'PhoneService', 'MultipleLines', 'InternetService',
                'OnlineSecurity', 'OnlineBackup', 'DeviceProtection',
                'TechSupport', 'StreamingTV', 'StreamingMovies',
                'Contract', 'PaperlessBilling', 'PaymentMethod',
                'MonthlyCharges', 'TotalCharges'
            ]

            missing = [col for col in required_cols if col not in df.columns]

            if missing:
                st.error(f"❌ Missing columns: {missing}")
                return

            # ------------------ PREDICTION ------------------
            results = []
            progress_bar = st.progress(0)

            for i, row in df.iterrows():

                input_data = row.to_dict()

                try:
                    processed = preprocess_input(input_data, label_encoders, scaler, features)
                    pred, prob = make_prediction(model, processed)

                    prob = round(prob * 100, 2)

                    if prob < 40:
                        risk = "Low"
                    elif prob < 70:
                        risk = "Medium"
                    else:
                        risk = "High"

                    input_data["Churn Probability (%)"] = prob
                    input_data["Risk Level"] = risk
                    input_data["Prediction"] = "Churn" if pred == 1 else "No Churn"

                except Exception as e:
                    input_data["Error"] = str(e)

                results.append(input_data)
                progress_bar.progress((i + 1) / len(df))

            result_df = pd.DataFrame(results)

            # ------------------ FILTERS ------------------
            st.subheader("🎛️ Filter Results")

            col1, col2 = st.columns(2)

            with col1:
                prediction_filter = st.selectbox(
                    "Filter by Prediction",
                    ["All", "Churn", "No Churn"]
                )

            with col2:
                risk_filter = st.selectbox(
                    "Filter by Risk Level",
                    ["All", "Low", "Medium", "High"]
                )

            # ------------------ APPLY FILTERS ------------------
            filtered_df = result_df.copy()

            if prediction_filter != "All":
                filtered_df = filtered_df[filtered_df["Prediction"] == prediction_filter]

            if risk_filter != "All":
                filtered_df = filtered_df[filtered_df["Risk Level"] == risk_filter]

            st.caption(f"Showing {len(filtered_df)} out of {len(result_df)} records")

            # ------------------ ROW COLOR FUNCTION ------------------
            def highlight_rows(row):
                risk = row.get("Risk Level", "")
                prediction = row.get("Prediction", "")

                if risk == "Low":
                    return ['background-color: #0bb80e'] * len(row)

                elif risk == "Medium" and prediction == "No Churn":
                    return ['background-color: #cfb913'] * len(row)

                elif risk == "Medium" and prediction == "Churn":
                    return ['background-color: #b36710'] * len(row)

                elif risk == "High":
                    return ['background-color: #b01322'] * len(row)

                else:
                    return [''] * len(row)

            styled_df = filtered_df.style.apply(highlight_rows, axis=1)

            # ------------------ OUTPUT ------------------
            st.subheader("📊 Prediction Results")
            st.dataframe(styled_df)

            # ------------------ DOWNLOAD CSV ------------------
            csv = result_df.to_csv(index=False).encode('utf-8')

            st.download_button(
                label="📥 Download CSV",
                data=csv,
                file_name="churn_predictions.csv",
                mime="text/csv"
            )

            # ------------------ STYLED EXCEL ------------------
            def generate_excel(df):

                output = BytesIO()

                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Predictions')

                    sheet = writer.sheets['Predictions']
                    from openpyxl.styles import PatternFill

                    for row in range(2, len(df) + 2):
                        risk = sheet[f"U{row}"].value
                        prediction = sheet[f"V{row}"].value

                        if risk == "Low":
                            fill = PatternFill(start_color="0BB80E", fill_type="solid")

                        elif risk == "Medium" and prediction == "No Churn":
                            fill = PatternFill(start_color="CFB913", fill_type="solid")

                        elif risk == "Medium" and prediction == "Churn":
                            fill = PatternFill(start_color="B36710", fill_type="solid")

                        elif risk == "High":
                            fill = PatternFill(start_color="B01322", fill_type="solid")

                        else:
                            fill = None

                        if fill:
                            for col in range(1, len(df.columns) + 1):
                                sheet.cell(row=row, column=col).fill = fill

                return output.getvalue()

            excel_data = generate_excel(result_df)

            st.download_button(
                label="📥 Download Colored Excel",
                data=excel_data,
                file_name="churn_predictions.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            # ------------------ SUMMARY ------------------
            st.subheader("📈 Summary")

            col1, col2, col3 = st.columns(3)

            churn_count = (result_df["Prediction"] == "Churn").sum()
            total = len(result_df)

            with col1:
                st.metric("Total Customers", total)

            with col2:
                st.metric("Churn Predicted", churn_count)

            with col3:
                st.metric("Churn Rate", f"{(churn_count/total)*100:.2f}%")

            # ------------------ SMART SUGGESTIONS ------------------
            st.subheader("💡 Business Recommendations")

            high = (result_df["Risk Level"] == "High").sum()
            medium = (result_df["Risk Level"] == "Medium").sum()
            low = (result_df["Risk Level"] == "Low").sum()

            if low > 0:
                st.markdown("### 🟢 Low Risk Customers")
                st.write("- Offer loyalty rewards or thank-you discount.")
                st.write("- Encourage the customer to leave a review or referral.")

            if medium > 0:
                st.markdown("### 🟡 Medium Risk Customers")
                st.write("- Send personalized offers or service upgrades.")
                st.write("- Reach out for feedback and experience improvement.")

            if high > 0:
                st.markdown("### 🔴 High Risk Customers")
                st.write("- Assign a retention specialist to the account.")
                st.write("- Offer strong incentives or extended contract discounts.")
                st.write("- Evaluate any service complaints urgently.")

        except Exception as e:
            st.error(f"❌ Error processing file: {e}")